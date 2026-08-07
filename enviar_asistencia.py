#!/usr/bin/env python3
"""
Automatizador de envío de asistencia a Google Forms
Ensamble de música - Proyectos Estudiantiles PGP 2026

Uso:
    python3 enviar_asistencia.py <URL_FORMS>

Ejemplo:
    python3 enviar_asistencia.py "https://docs.google.com/forms/d/e/1FAIpQL.../viewform"
"""

import sys
import os
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options as ChromeOptions
import time

# Constantes
PROYECTO = "Ensamble The Crumbs"
TIPO_USUARIO = "Estudiante de pregrado"
EXCEL_FILE = "formato.xlsx"
LOG_FILE = "asistencia_log.txt"

class ValidadorDatos:
    """Valida datos del Excel antes de enviar"""
    
    @staticmethod
    def validar_fecha(fecha_str):
        """Valida que la fecha sea válida. Formato: DD/MM"""
        if not fecha_str:
            return False, "Fecha vacía"
        
        # Intenta parsear diferentes formatos
        formatos = ["%d/%m", "%d/%m/%Y", "%Y-%m-%d"]
        for fmt in formatos:
            try:
                datetime.strptime(str(fecha_str).strip(), fmt)
                return True, None
            except:
                continue
        
        return False, f"Formato de fecha inválido: {fecha_str}"
    
    @staticmethod
    def validar_correo(correo):
        """Valida que el correo tenga formato @unal.edu.co"""
        if not correo:
            return False, "Correo vacío"
        
        correo = str(correo).strip()
        if not re.match(r"^[\w\.-]+@unal\.edu\.co$", correo):
            return False, f"Correo inválido (debe ser @unal.edu.co): {correo}"
        
        return True, None
    
    @staticmethod
    def validar_documento(doc):
        """Valida que el documento sea numérico"""
        if not doc:
            return False, "Documento vacío"
        
        if not str(doc).strip().isdigit():
            return False, f"Documento debe ser numérico: {doc}"
        
        return True, None
    
    @staticmethod
    def validar_texto(texto, campo_nombre):
        """Valida que el campo de texto no esté vacío"""
        if not texto or (isinstance(texto, str) and not texto.strip()):
            return False, f"{campo_nombre} vacío"
        return True, None

class LectorExcel:
    """Lee y procesa el archivo Excel"""
    
    def __init__(self, ruta_excel):
        if not Path(ruta_excel).exists():
            raise FileNotFoundError(f"No se encontró: {ruta_excel}")
        
        self.ruta = ruta_excel
        self.datos = []
        self.nombre_estudiante = None
        self.leer()
    
    def leer(self):
        """Lee el Excel y extrae datos válidos"""
        wb = load_workbook(self.ruta)
        ws = wb.active
        
        # Lee el nombre del estudiante de E2
        nombre_cell = ws['E2']
        if not nombre_cell.value:
            raise ValueError("❌ La casilla E2 está vacía. Debe contener el nombre del estudiante.")
        
        self.nombre_estudiante = str(nombre_cell.value).strip()
        
        # Encuentra headers en fila 2
        headers = {}
        for col_idx, cell in enumerate(ws[2], 1):
            if cell.value:
                headers[cell.value.strip()] = col_idx
        
        if not all(h in headers for h in ["Fecha", "Lugar", "Nombre del Evento", "Correo"]):
            raise ValueError("Excel debe tener columnas: Fecha, Lugar, Nombre del Evento, Correo")
        
        # Obtiene las posiciones de las columnas
        col_fecha = headers["Fecha"]
        col_lugar = headers["Lugar"]
        col_evento = headers["Nombre del Evento"]
        col_asistencia = 5  # Columna E siempre es la de asistencia
        col_correo = headers["Correo"]
        col_documento = headers.get("Documento", None)
        
        # Lee datos desde fila 3
        for row_idx in range(3, ws.max_row + 1):
            fecha = ws.cell(row_idx, col_fecha).value
            lugar = ws.cell(row_idx, col_lugar).value
            evento = ws.cell(row_idx, col_evento).value
            asistio = ws.cell(row_idx, col_asistencia).value
            correo = ws.cell(row_idx, col_correo).value
            doc = ws.cell(row_idx, col_documento).value if col_documento else None
            
            # Solo procesa si asistió = "Sí"
            if asistio and str(asistio).strip().lower() == "si":
                self.datos.append({
                    "fecha": fecha,
                    "lugar": lugar,
                    "evento": evento,
                    "correo": correo,
                    "documento": doc,
                    "fila": row_idx
                })
    
    def obtener_nombre_estudiante(self):
        return self.nombre_estudiante
    
    def obtener_registros(self):
        return self.datos

class AutomatizadorForms:
    """Automatiza el envío al Google Forms con Selenium"""
    
    def __init__(self, url_forms):
        self.url_forms = url_forms
        self.driver = None
        self.registros_enviados = 0
        self.registros_fallidos = 0
        self.errores = []
        self._iniciar_driver()
    
    def _iniciar_driver(self):
        """Inicia el driver de Chrome en modo headless"""
        chrome_options = ChromeOptions()
        #chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
        except Exception as e:
            print(f"❌ Error al iniciar Chrome: {e}")
            print("   Asegúrate de tener Chrome instalado y chromedriver disponible")
            print("   Instala: pip install webdriver-manager")
            sys.exit(1)
    
    def _formato_fecha(self, fecha):
        """Convierte fecha a formato DD/MM/YYYY para el formulario"""
        if not fecha:
            return ""
        
        fecha_str = str(fecha).strip()
        
        # Si ya es DD/MM, agregua año
        if re.match(r"^\d{2}/\d{2}$", fecha_str):
            return fecha_str + "/2026"
        
        # Si es DD/MM/YYYY, devuelve tal cual
        if re.match(r"^\d{2}/\d{2}/\d{4}$", fecha_str):
            return fecha_str
        
        # Si es YYYY-MM-DD (de Excel), convierte
        try:
            dt = datetime.strptime(fecha_str, "%Y-%m-%d")
            return dt.strftime("%d/%m/%Y")
        except:
            pass
        
        return fecha_str
    
    def _esperar_elemento(self, by, value, timeout=10):
        """Espera a que un elemento esté presente"""
        try:
            return WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located((by, value))
            )
        except Exception as e:
            raise TimeoutError(f"Elemento no encontrado ({by}={value}): {e}")
    
    def _llenar_campo_texto(self, campo_label, valor):
        """Llena un campo de texto del formulario"""
        try:
            # Busca el input asociado a este label
            inputs = self.driver.find_elements(By.TAG_NAME, "input")
            for inp in inputs:
                if inp.get_attribute("type") in ["text", "email"]:
                    inp.clear()
                    inp.send_keys(valor)
                    return True
            
            # Alternativa: busca textareas
            textareas = self.driver.find_elements(By.TAG_NAME, "textarea")
            for ta in textareas:
                ta.clear()
                ta.send_keys(valor)
                return True
            
            return False
        except Exception as e:
            raise Exception(f"Error llenando '{campo_label}': {e}")
    
    def _llenar_dropdown(self, valor_esperado):
        """Selecciona un valor en un dropdown"""
        try:
            selects = self.driver.find_elements(By.TAG_NAME, "select")
            for select in selects:
                try:
                    select_obj = Select(select)
                    select_obj.select_by_value(valor_esperado)
                    return True
                except:
                    continue
            
            # Si no hay select HTML, intenta con material select
            options = self.driver.find_elements(By.TAG_NAME, "div")
            for opt in options:
                if valor_esperado in opt.text:
                    opt.click()
                    return True
            
            return False
        except Exception as e:
            raise Exception(f"Error en dropdown: {e}")
    
    def enviar_registro(self, registro):
        """Envía un registro al formulario"""
        import traceback
        try:
            # Abre el formulario
            self.driver.get(self.url_forms)
            time.sleep(2)

            # Busca el campo de correo
            correos = self.driver.find_elements(By.CSS_SELECTOR, "input[aria-label*='correo'], input[aria-label*='email']")
            print(f"    🔍 Campos de correo encontrados: {len(correos)}")

            if correos:
                correos[0].clear()
                correos[0].send_keys(registro["correo"])
            else:
                print(f"    ⚠️  No se encontró campo de correo")

            # Busca botón "Siguiente" o similar y avanza
            botones = self.driver.find_elements(By.TAG_NAME, "button")
            print(f"    🔍 Botones encontrados: {len(botones)} -> {[b.text for b in botones]}")

            siguiente_encontrado = False
            for btn in botones:
                if "siguiente" in btn.text.lower() or "next" in btn.text.lower():
                    btn.click()
                    siguiente_encontrado = True
                    time.sleep(1.5)
                    break
                
            if not siguiente_encontrado:
                print(f"    ⚠️  No se encontró botón 'Siguiente'")

            # Busca campos para nombre del evento
            inputs_texto = self.driver.find_elements(By.CSS_SELECTOR, "input[type='text'], textarea")
            print(f"    🔍 Inputs de texto encontrados: {len(inputs_texto)}")
            idx = 0

            # ... (resto del método igual)

            # Envía el formulario
            botones_envio = self.driver.find_elements(By.TAG_NAME, "button")
            print(f"    🔍 Botones antes de enviar: {[b.text for b in botones_envio]}")

            for btn in botones_envio:
                if "enviar" in btn.text.lower() or "submit" in btn.text.lower():
                    btn.click()
                    time.sleep(2)
                    return True

            print(f"    ⚠️  No se encontró botón 'Enviar'")
            return False

        except Exception as e:
            print(f"    💥 EXCEPCIÓN: {type(e).__name__}: {e}")
            traceback.print_exc()
            # Screenshot para ver exactamente qué estaba en pantalla al fallar
            try:
                screenshot_path = f"error_fila_{registro['fila']}.png"
                self.driver.save_screenshot(screenshot_path)
                print(f"    📸 Screenshot guardado: {screenshot_path}")
            except:
                pass
            self.errores.append(f"Fila {registro['fila']}: {str(e)}")
            return False
    
    def procesar_lote(self, registros, nombre_estudiante):
        """Procesa un lote de registros"""
        print(f"\n📋 Procesando {len(registros)} registros...\n")
        
        for i, registro in enumerate(registros, 1):
            print(f"[{i}/{len(registros)}] Enviando: {registro['evento']} ({registro['fecha']})...")
            
            if self.enviar_registro(registro):
                self.registros_enviados += 1
                print(f"    ✅ Enviado exitosamente")
            else:
                self.registros_fallidos += 1
                print(f"    ❌ Error al enviar")
            
            # Pequeña pausa entre registros
            time.sleep(1)
        
        self._generar_log(nombre_estudiante)
        self.cerrar()
    
    def _generar_log(self, nombre_estudiante):
        """Genera archivo de log con resultados"""
        with open(LOG_FILE, "w", encoding="utf-8") as f:
            f.write("=" * 60 + "\n")
            f.write("LOG DE ENVÍO - ASISTENCIA A REUNIONES\n")
            f.write(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")
            
            f.write(f"Estudiante: {nombre_estudiante}\n")
            f.write(f"Proyecto: {PROYECTO}\n")
            f.write(f"Tipo de usuario: {TIPO_USUARIO}\n\n")
            
            f.write(f"Registros enviados: {self.registros_enviados}\n")
            f.write(f"Registros fallidos: {self.registros_fallidos}\n")
            f.write(f"Total procesado: {self.registros_enviados + self.registros_fallidos}\n\n")
            
            if self.errores:
                f.write("ERRORES:\n")
                for error in self.errores:
                    f.write(f"  - {error}\n")
            else:
                f.write("✅ No se registraron errores\n")
        
        print(f"\n📄 Log guardado en: {LOG_FILE}")
    
    def cerrar(self):
        """Cierra el navegador"""
        if self.driver:
            self.driver.quit()

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 enviar_asistencia.py <URL_FORMS>")
        print("\nEjemplo:")
        print("  python3 enviar_asistencia.py 'https://docs.google.com/forms/d/e/...'")
        sys.exit(1)
    
    url_forms = sys.argv[1]
    
    print("\n" + "=" * 60)
    print("🎵 AUTOMATIZADOR DE ASISTENCIA - ENSAMBLE THE CRUMBS 🎵")
    print("=" * 60 + "\n")
    
    try:
        # Paso 1: Leer Excel
        print("📁 Leyendo Excel...")
        lector = LectorExcel(EXCEL_FILE)
        nombre_estudiante = lector.obtener_nombre_estudiante()
        registros = lector.obtener_registros()
        
        print(f"👤 Estudiante: {nombre_estudiante}")
        
        if not registros:
            print(f"❌ No hay registros con asistencia (Sí) para {nombre_estudiante}")
            sys.exit(1)
        
        print(f"✅ Se encontraron {len(registros)} registros de asistencia")
        
        # Paso 2: Validar datos
        print("\n🔍 Validando datos...")
        validador = ValidadorDatos()
        
        for registro in registros:
            # Valida fecha
            valido, msg = validador.validar_fecha(registro["fecha"])
            if not valido:
                print(f"❌ Fila {registro['fila']}: {msg}")
                sys.exit(1)
            
            # Valida correo
            valido, msg = validador.validar_correo(registro["correo"])
            if not valido:
                print(f"❌ Fila {registro['fila']}: {msg}")
                sys.exit(1)
            
            # Valida documento
            if registro.get("documento"):
                valido, msg = validador.validar_documento(registro["documento"])
                if not valido:
                    print(f"❌ Fila {registro['fila']}: {msg}")
                    sys.exit(1)
            else:
                print(f"⚠️  Fila {registro['fila']}: Documento vacío (se enviará sin él)")
            
            # Valida textos
            valido, msg = validador.validar_texto(registro["evento"], "Nombre del Evento")
            if not valido:
                print(f"❌ Fila {registro['fila']}: {msg}")
                sys.exit(1)
            
            valido, msg = validador.validar_texto(registro["lugar"], "Lugar")
            if not valido:
                print(f"❌ Fila {registro['fila']}: {msg}")
                sys.exit(1)
        
        print("✅ Todos los datos son válidos\n")
        
        # Paso 3: Procesar con Selenium
        print(f"🌐 Abriendo Forms: {url_forms[:60]}...\n")
        automatizador = AutomatizadorForms(url_forms)
        automatizador.procesar_lote(registros, nombre_estudiante)
        
        # Resumen final
        print("\n" + "=" * 60)
        print("RESUMEN FINAL")
        print("=" * 60)
        print(f"👤 Estudiante: {nombre_estudiante}")
        print(f"✅ Enviados: {automatizador.registros_enviados}")
        print(f"❌ Fallidos: {automatizador.registros_fallidos}")
        print(f"📄 Log: {LOG_FILE}")
        print("=" * 60 + "\n")
    
    except FileNotFoundError as e:
        print(f"❌ {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"❌ {e}")
        sys.exit(1)
    except KeyboardInterrupt:
        print("\n\n⏹️  Proceso cancelado por el usuario")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Error inesperado: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
